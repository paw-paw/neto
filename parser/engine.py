"""Deterministic linear-table XLSX parser."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from typing import Any
from zipfile import BadZipFile
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH
from openpyxl.utils.exceptions import InvalidFileException

from .datetime_utils import (
    NonexistentLocalTimeError,
    normalize_whitespace,
    parse_date_value,
    parse_time_value,
    to_utc_string,
)
from .models import ParseResult, ParsedMatch, ParserKey, ValidationIssue
from .parser_keys import FIELDS, TEAM_FIELDS
from .validation import finalize_parse_result
from .workbook_security import WorkbookSafetyError, validate_xlsx_archive


EMPTY_ROW_LIMIT = 5


@dataclass(frozen=True)
class _CellValue:
    value: Any = None
    formula_without_cache: bool = False
    formula_text: str | None = None


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return normalize_whitespace(value) == ""
    return False


def _cell_is_empty(cell: _CellValue) -> bool:
    return not cell.formula_without_cache and _is_empty_value(cell.value)


def _normalized_text(cell: _CellValue) -> str:
    if cell.formula_without_cache or cell.value is None:
        return ""
    value = cell.value
    if isinstance(value, str):
        return normalize_whitespace(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date_display(cell: _CellValue) -> str:
    if cell.formula_without_cache:
        return ""
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _normalized_text(cell)


def _time_display(cell: _CellValue) -> str:
    if cell.formula_without_cache:
        return ""
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None).isoformat(timespec="seconds")
    if isinstance(value, time):
        return value.replace(tzinfo=None).isoformat(timespec="seconds")
    return _normalized_text(cell)


def _issue(
    source_row: int,
    severity: str,
    code: str,
    field: str,
    message: str,
) -> ValidationIssue:
    return ValidationIssue(
        source_row=source_row,
        severity=severity,
        code=code,
        affected_field=field,
        message=message,
    )


def _read_cell(value_sheet: Any, formula_sheet: Any, coordinate: str) -> _CellValue:
    value_cell = value_sheet[coordinate]
    formula_cell = formula_sheet[coordinate]
    is_formula = formula_cell.data_type == "f"
    formula_text = str(formula_cell.value) if is_formula else None
    return _CellValue(
        value=value_cell.value,
        formula_without_cache=is_formula and value_cell.value is None,
        formula_text=formula_text,
    )


def _read_mapped_row(
    value_sheet: Any, formula_sheet: Any, parser_key: ParserKey, row_number: int
) -> dict[str, _CellValue]:
    row: dict[str, _CellValue] = {}
    for field_name in FIELDS:
        column = parser_key.field_mappings.get(field_name)
        row[field_name] = (
            _read_cell(value_sheet, formula_sheet, f"{column}{row_number}")
            if column
            else _CellValue()
        )
    return row


def _row_is_empty(row: dict[str, _CellValue], parser_key: ParserKey) -> bool:
    mapped_fields = [
        field_name
        for field_name in FIELDS
        if parser_key.field_mappings.get(field_name) is not None
    ]
    return bool(mapped_fields) and all(_cell_is_empty(row[field]) for field in mapped_fields)


def _apply_forward_fill(
    raw_row: dict[str, _CellValue],
    carry: dict[str, _CellValue],
    parser_key: ParserKey,
) -> dict[str, _CellValue]:
    effective: dict[str, _CellValue] = {}
    for field_name in FIELDS:
        raw_cell = raw_row[field_name]
        can_fill = (
            field_name not in TEAM_FIELDS
            and parser_key.forward_fill_rules.get(field_name, False)
        )
        if _cell_is_empty(raw_cell) and can_fill and field_name in carry:
            effective[field_name] = carry[field_name]
        else:
            effective[field_name] = raw_cell

        if not raw_cell.formula_without_cache and not _cell_is_empty(raw_cell):
            carry[field_name] = raw_cell
    return effective


def _timezone_for_key(parser_key: ParserKey) -> tuple[ZoneInfo | None, str | None]:
    if not parser_key.base_timezone:
        return None, "missing_timezone"
    try:
        return ZoneInfo(parser_key.base_timezone), None
    except (ZoneInfoNotFoundError, ValueError):
        return None, "invalid_timezone"


def _parse_match(
    row_number: int,
    row: dict[str, _CellValue],
    parser_key: ParserKey,
    excel_epoch: datetime,
    zone: ZoneInfo | None,
    timezone_error: str | None,
) -> tuple[ParsedMatch, list[ValidationIssue]]:
    issues: list[ValidationIssue] = []
    parsed_date: date | None = None
    parsed_time: time | None = None

    date_cell = row["date"]
    time_cell = row["time"]

    if _cell_is_empty(date_cell):
        issues.append(
            _issue(row_number, "blocking_error", "missing_date", "date", "Date is required.")
        )
    elif date_cell.formula_without_cache:
        issues.append(
            _issue(
                row_number,
                "blocking_error",
                "unparseable_date",
                "date",
                "Date formula has no cached value.",
            )
        )
    else:
        try:
            parsed_date = parse_date_value(date_cell.value, epoch=excel_epoch)
        except (TypeError, ValueError, OverflowError):
            issues.append(
                _issue(
                    row_number,
                    "blocking_error",
                    "unparseable_date",
                    "date",
                    f'Could not parse date "{_date_display(date_cell)}".',
                )
            )

    if _cell_is_empty(time_cell):
        issues.append(
            _issue(row_number, "blocking_error", "missing_time", "time", "Time is required.")
        )
    elif time_cell.formula_without_cache:
        issues.append(
            _issue(
                row_number,
                "blocking_error",
                "unparseable_time",
                "time",
                "Time formula has no cached value.",
            )
        )
    else:
        try:
            parsed_time = parse_time_value(time_cell.value, epoch=excel_epoch)
        except (TypeError, ValueError, OverflowError):
            issues.append(
                _issue(
                    row_number,
                    "blocking_error",
                    "unparseable_time",
                    "time",
                    f'Could not parse time "{_time_display(time_cell)}".',
                )
            )

    if timezone_error == "missing_timezone":
        issues.append(
            _issue(
                row_number,
                "blocking_error",
                "missing_timezone",
                "timezone",
                "ParserKey timezone is required.",
            )
        )
    elif timezone_error == "invalid_timezone":
        issues.append(
            _issue(
                row_number,
                "blocking_error",
                "invalid_timezone",
                "timezone",
                f'ParserKey timezone "{parser_key.base_timezone}" is not a valid IANA timezone.',
            )
        )

    team_a = _normalized_text(row["team_a"])
    team_b = _normalized_text(row["team_b"])
    stage = _normalized_text(row["stage"])
    bo = _normalized_text(row["bo"])
    match_label = _normalized_text(row["match_label"])

    empty_participant_policy = parser_key.raw_data.get("validation_rules", {}).get(
        "empty_participant_policy", "blocking_error"
    )
    team_values = {"team_a": team_a, "team_b": team_b}
    for field_name in ("team_a", "team_b"):
        if team_values[field_name]:
            continue
        if empty_participant_policy == "use_tbd":
            team_values[field_name] = "TBD"
            if row[field_name].formula_without_cache:
                issues.append(
                    _issue(
                        row_number,
                        "warning",
                        "formula_cached_value_missing",
                        field_name,
                        "Participant formula has no cached value; explicit key policy used TBD.",
                    )
                )
        else:
            issues.append(
                _issue(
                    row_number,
                    "blocking_error",
                    f"missing_{field_name}",
                    field_name,
                    f"{field_name.replace('_', ' ').title()} is required.",
                )
            )
    team_a = team_values["team_a"]
    team_b = team_values["team_b"]

    for field_name, value, code, label in (
        ("stage", stage, "stage_missing", "Stage"),
        ("bo", bo, "bo_missing", "Best-of"),
        ("match_label", match_label, "match_label_missing", "Match label"),
    ):
        if not value:
            issues.append(
                _issue(
                    row_number,
                    "warning",
                    code,
                    field_name,
                    f"{label} is missing.",
                )
            )

    start_time_utc = ""
    if parsed_date is not None and parsed_time is not None and zone is not None:
        try:
            start_time_utc = to_utc_string(parsed_date, parsed_time, zone)
        except (NonexistentLocalTimeError, OverflowError, ValueError):
            issues.append(
                _issue(
                    row_number,
                    "blocking_error",
                    "unparseable_datetime",
                    "start_time_utc",
                    "The local date and time could not be converted to UTC.",
                )
            )

    match = ParsedMatch(
        source_row=row_number,
        date_original=_date_display(date_cell),
        time_original=_time_display(time_cell),
        timezone=parser_key.base_timezone,
        start_time_utc=start_time_utc,
        team_a=team_a,
        team_b=team_b,
        stage=stage,
        bo=bo,
        match_label=match_label,
    )
    return match, issues


def _timezone_notice(parser_key: ParserKey) -> str:
    value = parser_key.base_timezone or "(missing)"
    return f"timezone_from_key: match timezone is taken from ParserKey ({value})."


def parse_workbook(file_bytes: bytes, parser_key: ParserKey) -> ParseResult:
    """Parse an uploaded XLSX and return a complete, non-raising ParseResult."""

    try:
        validate_xlsx_archive(file_bytes)
    except WorkbookSafetyError as exc:
        return ParseResult.failed(str(exc))

    if parser_key.schema_version == "neto.parser_key.v2":
        from .v2_runtime import parse_workbook_v2

        return parse_workbook_v2(file_bytes, parser_key)

    notice = _timezone_notice(parser_key)
    value_workbook = None
    formula_workbook = None
    try:
        value_workbook = load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=False, keep_links=False
        )
        formula_workbook = load_workbook(
            BytesIO(file_bytes), data_only=False, read_only=False, keep_links=False
        )

        if parser_key.target_sheet not in value_workbook.sheetnames:
            return ParseResult.failed(
                f'The selected parser key expects sheet "{parser_key.target_sheet}", '
                "but it was not found in this file.",
                notice=notice,
            )

        value_sheet = value_workbook[parser_key.target_sheet]
        formula_sheet = formula_workbook[parser_key.target_sheet]
        excel_epoch = getattr(value_workbook, "epoch", WINDOWS_EPOCH)
        zone, timezone_error = _timezone_for_key(parser_key)

        matches: list[ParsedMatch] = []
        issues: list[ValidationIssue] = []
        carry: dict[str, _CellValue] = {}
        consecutive_empty = 0

        for row_number in range(parser_key.data_start_row, value_sheet.max_row + 1):
            raw_row = _read_mapped_row(
                value_sheet, formula_sheet, parser_key, row_number
            )
            if _row_is_empty(raw_row, parser_key):
                consecutive_empty += 1
                if consecutive_empty >= EMPTY_ROW_LIMIT:
                    break
                continue

            consecutive_empty = 0
            effective_row = _apply_forward_fill(raw_row, carry, parser_key)
            match, row_issues = _parse_match(
                row_number,
                effective_row,
                parser_key,
                excel_epoch,
                zone,
                timezone_error,
            )
            matches.append(match)
            issues.extend(row_issues)

        if not matches:
            return ParseResult.failed(
                "No non-empty mapped match rows were found in the target sheet.",
                notice=notice,
            )

        return finalize_parse_result(
            matches,
            issues,
            notice=notice,
        )
    except (BadZipFile, InvalidFileException, OSError, EOFError) as exc:
        return ParseResult.failed(
            f"Could not read this XLSX file. Upload a valid .xlsx schedule file. ({exc})",
            notice=notice,
        )
    except Exception as exc:  # Defensive UI boundary after expected failures are handled.
        return ParseResult.failed(
            f"The parser could not complete because of a technical error: {type(exc).__name__}: {exc}",
            notice=notice,
        )
    finally:
        if value_workbook is not None:
            value_workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()
