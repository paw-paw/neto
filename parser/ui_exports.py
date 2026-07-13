"""Serializers for the filtered match-table view."""

from __future__ import annotations

import html
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, TableStyle

from .models import OUTPUT_COLUMNS


DISPLAY_LABELS: dict[str, str] = {
    "date": "Date",
    "time": "Time",
    "team_a": "Team A",
    "team_b": "Team B",
    "bo": "BO",
    "stage": "Stage",
    "match_label": "Match",
    "timezone": "Schedule timezone",
    "start_time_utc": "Start time UTC",
    "row_status": "Status",
}


def canonical_csv_bytes(canonical: pd.DataFrame) -> bytes:
    """Serialize the filtered view with NETO's fixed canonical schema."""

    csv_text = canonical.loc[:, list(OUTPUT_COLUMNS)].to_csv(
        index=False, lineterminator="\n"
    )
    return csv_text.encode("utf-8-sig")


def _display_dataframe(presentation: pd.DataFrame) -> pd.DataFrame:
    return presentation.rename(columns=DISPLAY_LABELS)


def markdown_bytes(presentation: pd.DataFrame) -> bytes:
    """Create a dependency-free GitHub-flavored Markdown table."""

    dataframe = _display_dataframe(presentation)

    def clean(value: object) -> str:
        return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")

    header = "| " + " | ".join(map(clean, dataframe.columns)) + " |"
    separator = "| " + " | ".join("---" for _ in dataframe.columns) + " |"
    rows = [
        "| " + " | ".join(clean(value) for value in row) + " |"
        for row in dataframe.itertuples(index=False, name=None)
    ]
    return ("\n".join([header, separator, *rows]) + "\n").encode("utf-8")


def xlsx_bytes(presentation: pd.DataFrame) -> bytes:
    """Create a styled workbook of the filtered, human-readable view."""

    dataframe = _display_dataframe(presentation)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NETO Matches"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{max(1, len(dataframe) + 1)}"

    for column, value in enumerate(dataframe.columns, start=1):
        cell = sheet.cell(row=1, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24334B")
        cell.alignment = Alignment(vertical="center")

    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column, value="" if pd.isna(value) else str(value))

    widths = (13, 9, 24, 24, 8, 24, 20, 23, 24, 12)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def pdf_bytes(presentation: pd.DataFrame) -> bytes:
    """Create a landscape PDF table with repeating headers."""

    dataframe = _display_dataframe(presentation)
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="NETO Match Schedule",
    )
    styles = getSampleStyleSheet()
    cell_style = styles["BodyText"]
    cell_style.fontName = "Helvetica"
    cell_style.fontSize = 6.5
    cell_style.leading = 8

    table_rows = [
        [Paragraph(f"<b>{html.escape(str(column))}</b>", cell_style) for column in dataframe.columns]
    ]
    for row in dataframe.itertuples(index=False, name=None):
        table_rows.append(
            [
                Paragraph(html.escape("" if pd.isna(value) else str(value)), cell_style)
                for value in row
            ]
        )

    widths = [20, 14, 34, 34, 12, 31, 27, 29, 36, 18]
    table = LongTable(
        table_rows,
        colWidths=[width * mm for width in widths],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24334B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCD2DA")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FA")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    document.build([Paragraph("NETO Match Schedule", styles["Title"]), table])
    return buffer.getvalue()
