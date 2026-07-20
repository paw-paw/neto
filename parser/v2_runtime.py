"""Operator runtime for NETO ParserKey v2 documents."""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from typing import Any
from zipfile import BadZipFile
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
)
from openpyxl.utils.datetime import WINDOWS_EPOCH
from openpyxl.utils.exceptions import InvalidFileException

from .datetime_utils import (
    AmbiguousLocalTimeError,
    NonexistentLocalTimeError,
    normalize_whitespace,
    parse_date_value,
    parse_time_value,
    to_utc_string,
)
from .models import ParseResult, ParsedMatch, ParserKey, ValidationIssue


DIRECT_FORMULA_REFERENCE = re.compile(
    r"^=\s*(?:(?:'(?P<quoted_sheet>(?:[^']|'')+)'|"
    r"(?P<sheet>[A-Za-z_][A-Za-z0-9_.]*))!)?"
    r"\$?(?P<column>[A-Za-z]{1,3})\$?(?P<row>[1-9]\d*)\s*$"
)
MAX_FORMULA_REFERENCE_DEPTH = 8


@dataclass(frozen=True)
class _Anchor:
    row: int
    column: int
    tile_origin: tuple[int, int] | None = None
    tile_id: str | None = None


@dataclass
class _EvalValue:
    value: Any = None
    coordinate: str | None = None
    source_op: str | None = None
    transforms: list[str] = field(default_factory=list)
    used_fallback: bool = False
    used_default: bool = False
    override_reason: str | None = None
    missing_reason: str | None = None
    formula_cache_missing: bool = False
    formula_text: str | None = None
    formula_resolution: str | None = None
    resolved_source_cell: str | None = None
    fallback_reason: str | None = None
    source_value: Any = None

    def clone(self) -> "_EvalValue":
        return _EvalValue(
            value=self.value,
            coordinate=self.coordinate,
            source_op=self.source_op,
            transforms=list(self.transforms),
            used_fallback=self.used_fallback,
            used_default=self.used_default,
            override_reason=self.override_reason,
            missing_reason=self.missing_reason,
            formula_cache_missing=self.formula_cache_missing,
            formula_text=self.formula_text,
            formula_resolution=self.formula_resolution,
            resolved_source_cell=self.resolved_source_cell,
            fallback_reason=self.fallback_reason,
            source_value=self.source_value,
        )


@dataclass
class _Environment:
    runtime: "_Runtime"
    source_id: str
    record_set_id: str
    anchor: _Anchor
    context: dict[str, _EvalValue] = field(default_factory=dict)
    extractors: dict[str, dict[str, _EvalValue]] = field(default_factory=dict)


class _Runtime:
    def __init__(self, parser_key: ParserKey, value_workbook: Any, formula_workbook: Any):
        self.parser_key = parser_key
        self.config = parser_key.raw_data
        self.value_workbook = value_workbook
        self.formula_workbook = formula_workbook
        self.epoch = getattr(value_workbook, "epoch", WINDOWS_EPOCH)
        self.source_config = {
            source["source_id"]: source for source in self.config["sources"]
        }
        self.source_sheets: dict[str, tuple[Any, Any]] = {}

        for source_id, source in self.source_config.items():
            sheet_name = source["sheet_locator"]["args"]["sheet_name"]
            if sheet_name not in value_workbook.sheetnames:
                if source.get("required", True):
                    raise KeyError(
                        f'The selected parser key expects sheet "{sheet_name}", '
                        "but it was not found in this file."
                    )
                continue
            self.source_sheets[source_id] = (
                value_workbook[sheet_name],
                formula_workbook[sheet_name],
            )

    def sheets(self, source_id: str) -> tuple[Any, Any]:
        if source_id not in self.source_sheets:
            raise KeyError(f'ParserKey source "{source_id}" is unavailable.')
        return self.source_sheets[source_id]

    def sheet_name(self, source_id: str) -> str:
        return self.source_config[source_id]["sheet_locator"]["args"]["sheet_name"]

    @staticmethod
    def _formula_text(value: Any) -> str | None:
        if isinstance(value, str):
            return value
        text = getattr(value, "text", None)
        return text if isinstance(text, str) else None

    def _direct_formula_value(
        self,
        sheet_name: str,
        coordinate: str,
        *,
        depth: int = 0,
        visited: set[tuple[str, str]] | None = None,
    ) -> tuple[Any, str, str]:
        qualified_coordinate = f"{sheet_name}!{coordinate}"
        if depth >= MAX_FORMULA_REFERENCE_DEPTH:
            return None, "max_depth_exceeded", qualified_coordinate

        visited = set() if visited is None else set(visited)
        identity = (sheet_name.casefold(), coordinate.upper())
        if identity in visited:
            return None, "cycle_detected", qualified_coordinate
        visited.add(identity)

        if (
            sheet_name not in self.value_workbook.sheetnames
            or sheet_name not in self.formula_workbook.sheetnames
        ):
            return None, "sheet_not_found", qualified_coordinate

        cached_cell = self.value_workbook[sheet_name][coordinate]
        raw_cell = self.formula_workbook[sheet_name][coordinate]
        if cached_cell.value is not None:
            return cached_cell.value, "resolved_reference", qualified_coordinate
        if raw_cell.data_type != "f":
            if _is_empty(raw_cell.value):
                return None, "referenced_cell_empty", qualified_coordinate
            return raw_cell.value, "resolved_reference", qualified_coordinate

        formula_text = self._formula_text(raw_cell.value)
        match = (
            DIRECT_FORMULA_REFERENCE.fullmatch(formula_text or "")
            if isinstance(raw_cell.value, str)
            else None
        )
        if match is None:
            return None, "referenced_formula_not_direct", qualified_coordinate

        target_sheet = (
            match.group("quoted_sheet") or match.group("sheet") or sheet_name
        ).replace("''", "'")
        target_coordinate = f'{match.group("column").upper()}{match.group("row")}'
        return self._direct_formula_value(
            target_sheet,
            target_coordinate,
            depth=depth + 1,
            visited=visited,
        )

    def read_coordinate(self, source_id: str, coordinate: str) -> _EvalValue:
        value_sheet, formula_sheet = self.sheets(source_id)
        cached_cell = value_sheet[coordinate]
        raw_cell = formula_sheet[coordinate]
        is_formula = raw_cell.data_type == "f"
        if cached_cell.value is not None:
            return _EvalValue(
                value=cached_cell.value,
                coordinate=coordinate,
                source_op="cell",
                source_value=cached_cell.value,
            )
        if not is_formula:
            return _EvalValue(
                value=raw_cell.value,
                coordinate=coordinate,
                source_op="cell",
                missing_reason="source_empty" if _is_empty(raw_cell.value) else None,
                source_value=raw_cell.value,
            )

        formula_text = self._formula_text(raw_cell.value)
        reference = (
            DIRECT_FORMULA_REFERENCE.fullmatch(formula_text or "")
            if isinstance(raw_cell.value, str)
            else None
        )
        if reference is not None:
            current_sheet = self.sheet_name(source_id)
            target_sheet = (
                reference.group("quoted_sheet")
                or reference.group("sheet")
                or current_sheet
            ).replace("''", "'")
            target_coordinate = (
                f'{reference.group("column").upper()}{reference.group("row")}'
            )
            resolved, resolution, resolved_source_cell = self._direct_formula_value(
                target_sheet,
                target_coordinate,
            )
            if not _is_empty(resolved):
                return _EvalValue(
                    value=resolved,
                    coordinate=coordinate,
                    source_op="cell",
                    formula_cache_missing=True,
                    formula_text=formula_text,
                    formula_resolution=resolution,
                    resolved_source_cell=resolved_source_cell,
                    source_value=raw_cell.value,
                )
            return _EvalValue(
                value=None,
                coordinate=coordinate,
                source_op="cell",
                missing_reason="formula_cache_missing",
                formula_cache_missing=True,
                formula_text=formula_text,
                formula_resolution=resolution,
                resolved_source_cell=resolved_source_cell,
                source_value=raw_cell.value,
            )
        return _EvalValue(
            value=None,
            coordinate=coordinate,
            source_op="cell",
            missing_reason="formula_cache_missing",
            formula_cache_missing=True,
            formula_text=formula_text,
            formula_resolution=(
                "formula_not_direct_reference"
                if isinstance(raw_cell.value, str)
                else "array_formula_not_supported"
            ),
            source_value=raw_cell.value,
        )


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and normalize_whitespace(value) == "")


def _regex_flags(names: list[str] | None) -> int:
    result = 0
    for name in names or []:
        flag = getattr(re, name, None)
        if flag is None:
            raise ValueError(f'Unsupported regular-expression flag "{name}".')
        result |= flag
    return result


def _source_coordinate(column: int, row: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _merged_origin_coordinate(sheet: Any, coordinate: str) -> str:
    row, column = coordinate_to_tuple(coordinate)
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= column <= merged_range.max_col:
            return _source_coordinate(merged_range.min_col, merged_range.min_row)
    return coordinate


def _resolve_source(node: dict[str, Any], env: _Environment) -> _EvalValue:
    op = node["op"]
    args = node.get("args", {})
    runtime = env.runtime

    if op == "cell.column":
        coordinate = f'{args["column"]}{env.anchor.row}'
        value = runtime.read_coordinate(env.source_id, coordinate)
        value.source_op = op
        return value
    if op == "cell.absolute":
        value = runtime.read_coordinate(env.source_id, args["address"])
        value.source_op = op
        return value
    if op == "cell.relative":
        if args["base"] == "record_anchor":
            base_row, base_column = env.anchor.row, env.anchor.column
        elif args["base"] == "tile_origin" and env.anchor.tile_origin:
            base_row, base_column = env.anchor.tile_origin
        else:
            return _EvalValue(missing_reason="invalid_relative_base", source_op=op)
        coordinate = _source_coordinate(
            base_column + int(args.get("column_offset", 0)),
            base_row + int(args.get("row_offset", 0)),
        )
        value = runtime.read_coordinate(env.source_id, coordinate)
        value.source_op = op
        return value
    if op == "cell.merged_origin":
        value_sheet, _ = runtime.sheets(env.source_id)
        coordinate = f'{args["column"]}{env.anchor.row}'
        coordinate = _merged_origin_coordinate(value_sheet, coordinate)
        value = runtime.read_coordinate(env.source_id, coordinate)
        if _is_empty(value.value) and args.get("fallback") == "nearest_non_empty_above":
            current_row, column = coordinate_to_tuple(coordinate)
            for row_number in range(current_row - 1, 0, -1):
                candidate_coordinate = _source_coordinate(column, row_number)
                candidate_origin = _merged_origin_coordinate(
                    value_sheet, candidate_coordinate
                )
                candidate = runtime.read_coordinate(env.source_id, candidate_origin)
                if not _is_empty(candidate.value):
                    value = candidate
                    value.used_fallback = True
                    break
        value.source_op = op
        return value
    if op == "record.row_index":
        return _EvalValue(value=env.anchor.row, source_op=op)
    if op == "literal.value":
        return _EvalValue(value=args.get("value"), source_op=op)
    if op == "context.get":
        value = env.context.get(args["key"])
        return value.clone() if value else _EvalValue(source_op=op, missing_reason="source_empty")
    if op == "extractor.output":
        outputs = env.extractors.get(args["extractor_id"], {})
        value = outputs.get(args["output"])
        return value.clone() if value else _EvalValue(source_op=op, missing_reason="extractor_failed")
    raise ValueError(f'Unsupported value source "{op}".')


def _parse_date_with_formats(value: Any, formats: list[str], epoch: datetime) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if "auto" in formats:
        return parse_date_value(value, epoch=epoch)
    text = normalize_whitespace(str(value))
    for date_format in formats:
        candidates = [text]
        if "," in text:
            candidates.append(re.sub(r",\s*", ", ", text))
        for candidate in dict.fromkeys(candidates):
            parse_text = candidate
            parse_format = date_format
            if "%Y" not in date_format and "%y" not in date_format:
                parse_text = f"{candidate} 2000"
                parse_format = f"{date_format} %Y"
            try:
                return datetime.strptime(parse_text, parse_format).date()
            except ValueError:
                continue
    raise ValueError(f"Date does not match declared formats: {value!r}")


def _parse_time_with_formats(value: Any, formats: list[str], epoch: datetime) -> time:
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if "auto" in formats:
        return parse_time_value(value, epoch=epoch)
    text = normalize_whitespace(str(value)).upper()
    for time_format in formats:
        try:
            return datetime.strptime(text, time_format).time()
        except ValueError:
            continue
    raise ValueError(f"Time does not match declared formats: {value!r}")


def _apply_transform(
    current: _EvalValue, transform: dict[str, Any], epoch: datetime
) -> _EvalValue:
    result = current.clone()
    op = transform["op"]
    args = transform.get("args", {})
    result.transforms.append(op)
    if _is_empty(result.value):
        result.value = None
        result.missing_reason = result.missing_reason or "source_empty"
        return result

    try:
        if op == "text.to_string":
            if isinstance(result.value, datetime):
                result.value = result.value.isoformat(sep=" ")
            elif isinstance(result.value, (date, time)):
                result.value = result.value.isoformat()
            else:
                result.value = str(result.value)
        elif op == "text.trim":
            result.value = str(result.value).strip()
        elif op == "text.regex_extract":
            match = re.search(
                args["pattern"], str(result.value), _regex_flags(args.get("flags"))
            )
            result.value = match.group(args.get("group", 0)) if match else None
        elif op == "text.regex_replace":
            result.value = re.sub(
                args["pattern"],
                args.get("replacement", ""),
                str(result.value),
                flags=_regex_flags(args.get("flags")),
            )
        elif op == "datetime.excel_serial_to_date":
            result.value = parse_date_value(result.value, epoch=epoch)
        elif op == "datetime.excel_fraction_to_time":
            parsed = parse_time_value(result.value, epoch=epoch)
            result.value = parsed.strftime(args.get("output_format", "%H:%M:%S"))
        elif op == "datetime.parse_date":
            result.value = _parse_date_with_formats(
                result.value, args.get("formats", ["auto"]), epoch
            )
        elif op == "datetime.parse_time":
            result.value = _parse_time_with_formats(
                result.value, args.get("formats", ["auto"]), epoch
            )
        elif op == "datetime.inject_year":
            if isinstance(result.value, datetime):
                result.value = result.value.replace(year=int(args["year"]))
            elif isinstance(result.value, date):
                result.value = result.value.replace(year=int(args["year"]))
            else:
                raise ValueError("inject_year requires a parsed date")
        elif op == "datetime.format_date":
            parsed_date = (
                result.value.date()
                if isinstance(result.value, datetime)
                else result.value
            )
            if not isinstance(parsed_date, date):
                parsed_date = parse_date_value(parsed_date, epoch=epoch)
            result.value = parsed_date.strftime(args["format"])
        elif op == "datetime.format_time":
            parsed_time = (
                result.value.time()
                if isinstance(result.value, datetime)
                else result.value
            )
            if not isinstance(parsed_time, time):
                parsed_time = parse_time_value(parsed_time, epoch=epoch)
            result.value = parsed_time.strftime(args["format"])
        else:
            raise ValueError(f'Unsupported transform "{op}".')
    except (TypeError, ValueError, OverflowError, re.error):
        result.value = None
        result.missing_reason = "transform_failed"
    return result


def _evaluate_pipeline(
    pipeline: dict[str, Any], env: _Environment, epoch: datetime
) -> _EvalValue:
    result = _resolve_source(pipeline["source"], env)
    for transform in pipeline.get("transforms", []):
        result = _apply_transform(result, transform, epoch)
        if result.value is None:
            break
    if not _is_empty(result.value):
        return result

    for fallback in pipeline.get("fallbacks", []):
        candidate = _evaluate_pipeline(fallback, env, epoch)
        if not _is_empty(candidate.value):
            candidate.used_fallback = True
            candidate.fallback_reason = result.missing_reason or "source_empty"
            if result.formula_cache_missing:
                candidate.formula_cache_missing = True
                candidate.formula_text = result.formula_text
                candidate.formula_resolution = result.formula_resolution
                candidate.resolved_source_cell = result.resolved_source_cell
                candidate.coordinate = result.coordinate
                candidate.source_op = result.source_op
                candidate.source_value = result.source_value
            return candidate
    result.value = None
    result.missing_reason = result.missing_reason or "source_empty"
    return result


def _predicate_operand(value: Any, env: _Environment) -> Any:
    if isinstance(value, dict) and "op" in value:
        return _resolve_source(value, env).value
    return value


def _evaluate_predicate(predicate: dict[str, Any] | None, env: _Environment) -> bool:
    if predicate is None:
        return True
    op = predicate["op"]
    args = predicate.get("args", {})
    if op == "predicate.all":
        return all(_evaluate_predicate(item, env) for item in args["predicates"])
    if op == "predicate.any":
        return any(_evaluate_predicate(item, env) for item in args["predicates"])
    if op == "predicate.not_empty":
        return not _is_empty(_predicate_operand(args["value"], env))
    if op == "predicate.equals":
        return _predicate_operand(args["left"], env) == _predicate_operand(
            args["right"], env
        )
    if op == "predicate.matches_regex":
        value = _predicate_operand(args["value"], env)
        if _is_empty(value):
            return False
        return (
            re.search(
                args["pattern"], str(value), _regex_flags(args.get("flags"))
            )
            is not None
        )
    raise ValueError(f'Unsupported predicate "{op}".')


def _anchors(locator: dict[str, Any]) -> list[_Anchor]:
    op = locator["op"]
    args = locator.get("args", {})
    anchors: list[_Anchor] = []
    if op == "records.row_ranges":
        column = column_index_from_string(args["anchor_column"])
        for row_range in args["ranges"]:
            anchors.extend(
                _Anchor(row=row_number, column=column)
                for row_number in range(
                    int(row_range["start_row"]), int(row_range["end_row"]) + 1
                )
            )
        return anchors
    if op == "records.row_scan":
        column = column_index_from_string(args["anchor_column"])
        return [
            _Anchor(row=row_number, column=column)
            for row_number in range(
                int(args["start_row"]),
                int(args["end_row"]) + 1,
                int(args.get("step", 1)),
            )
        ]
    if op == "records.tile_grid":
        column_offset = int(args.get("record_anchor", {}).get("column_offset", 0))
        for tile in args["tiles"]:
            origin_row, origin_column = coordinate_to_tuple(tile["origin"])
            for row_offset in tile["record_row_offsets"]:
                anchors.append(
                    _Anchor(
                        row=origin_row + int(row_offset),
                        column=origin_column + column_offset,
                        tile_origin=(origin_row, origin_column),
                        tile_id=tile["tile_id"],
                    )
                )
        return anchors
    raise ValueError(f'Unsupported record locator "{op}".')


def _extractor_outputs(
    extractor: dict[str, Any], env: _Environment, epoch: datetime
) -> dict[str, _EvalValue]:
    input_value = _evaluate_pipeline(extractor["input"], env, epoch)
    operator = extractor["operator"]
    if operator["op"] != "extract.regex" or _is_empty(input_value.value):
        return {}

    args = operator["args"]
    for pattern_spec in args["patterns"]:
        match = re.search(
            pattern_spec["pattern"],
            str(input_value.value),
            _regex_flags(pattern_spec.get("flags")),
        )
        if not match:
            continue
        outputs: dict[str, _EvalValue] = {}
        for output_name in args["named_outputs"]:
            output = input_value.clone()
            output.value = match.groupdict().get(output_name)
            output.source_op = "extract.regex"
            output.missing_reason = None if output.value is not None else "extractor_failed"
            outputs[output_name] = output
        return outputs
    return {
        output_name: _failed_extractor_output(input_value)
        for output_name in args["named_outputs"]
    }


def _failed_extractor_output(input_value: _EvalValue) -> _EvalValue:
    output = input_value.clone()
    output.value = None
    output.source_op = "extract.regex"
    output.missing_reason = "extractor_failed"
    return output


def _field_value(
    field_spec: dict[str, Any], env: _Environment, epoch: datetime
) -> _EvalValue:
    result = _evaluate_pipeline(field_spec["value"], env, epoch)
    if _is_empty(result.value) and field_spec.get("on_missing") == "use_default":
        result.fallback_reason = result.missing_reason or "source_empty"
        result.value = field_spec.get("default")
        result.used_default = True
        result.missing_reason = None if not _is_empty(result.value) else result.missing_reason
    return result


def _normalized_output(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str) and value.lstrip().startswith("="):
        return ""
    if not isinstance(value, (str, int, float, bool, date, time)):
        return ""
    if isinstance(value, datetime):
        value = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        value = value.isoformat()
    return normalize_whitespace(str(value))


def _normalize_bo(value: str, enabled: bool) -> str:
    if not enabled or not value:
        return value
    match = re.search(r"(?i)\b(?:bo|best\s*of)\s*[-:]?\s*(\d+)\b", value)
    return f"BO{match.group(1)}" if match else value


def _issue_for_missing(
    field_name: str,
    result: _EvalValue,
    field_spec: dict[str, Any],
    env: _Environment,
) -> ValidationIssue | None:
    source_text = _normalized_output(result.source_value).upper()
    if field_name == "time" and source_text in {"TBD", "TBA"}:
        return ValidationIssue(
            source_row=env.anchor.row,
            source_sheet=env.runtime.sheet_name(env.source_id),
            record_set_id=env.record_set_id,
            severity="warning",
            code="time_pending",
            affected_field=field_name,
            message="Time is pending in the source schedule.",
        )
    on_missing = field_spec.get("on_missing", "null")
    if on_missing not in {"blocking_error", "warning"}:
        return None
    severity = "blocking_error" if on_missing == "blocking_error" else "warning"
    if field_name == "date":
        code = "unparseable_date" if result.missing_reason == "transform_failed" else "missing_date"
    elif field_name == "time":
        code = "unparseable_time" if result.missing_reason == "transform_failed" else "missing_time"
    elif field_name in {"team_a", "team_b"}:
        code = f"missing_{field_name}"
    else:
        code = f"{field_name}_missing"
    return ValidationIssue(
        source_row=env.anchor.row,
        source_sheet=env.runtime.sheet_name(env.source_id),
        record_set_id=env.record_set_id,
        severity=severity,
        code=code,
        affected_field=field_name,
        message=f'{field_name.replace("_", " ").title()} is missing or could not be parsed.',
    )


def _provenance(result: _EvalValue, env: _Environment) -> dict[str, Any]:
    return {
        "source_workbook": env.runtime.parser_key.raw_data.get("metadata", {}).get(
            "source_files", [{}]
        )[0].get("filename"),
        "source_sheet": env.runtime.sheet_name(env.source_id),
        "source_cell": result.coordinate,
        "record_set": env.record_set_id,
        "tile_id": env.anchor.tile_id,
        "tile_origin": (
            _source_coordinate(env.anchor.tile_origin[1], env.anchor.tile_origin[0])
            if env.anchor.tile_origin
            else None
        ),
        "source_operator": result.source_op,
        "transform_chain": list(result.transforms),
        "used_fallback": result.used_fallback,
        "used_default": result.used_default,
        "override_reason": result.override_reason,
        "formula_cache_missing": result.formula_cache_missing,
        "formula_text": result.formula_text,
        "formula_resolution": result.formula_resolution,
        "resolved_source_cell": result.resolved_source_cell,
        "fallback_reason": result.fallback_reason,
    }


def _parse_record(
    runtime: _Runtime,
    record_set: dict[str, Any],
    anchor: _Anchor,
    context: dict[str, _EvalValue],
    zone: ZoneInfo | None,
    timezone_error: str | None,
) -> tuple[ParsedMatch | None, list[ValidationIssue]]:
    env = _Environment(
        runtime=runtime,
        source_id=record_set["source_id"],
        record_set_id=record_set["record_set_id"],
        anchor=anchor,
        context=context,
    )
    if not _evaluate_predicate(record_set.get("filter"), env):
        return None, []

    for extractor in record_set.get("extractors", []):
        env.extractors[extractor["extractor_id"]] = _extractor_outputs(
            extractor, env, runtime.epoch
        )

    field_specs = dict(record_set["fields"])
    values = {
        field_name: _field_value(field_spec, env, runtime.epoch)
        for field_name, field_spec in field_specs.items()
    }

    for override in record_set.get("overrides", []):
        if not _evaluate_predicate(override["when"], env):
            continue
        for field_name, field_spec in override["set"].items():
            field_specs[field_name] = field_spec
            value = _field_value(field_spec, env, runtime.epoch)
            value.override_reason = override.get("reason")
            values[field_name] = value

    issues: list[ValidationIssue] = []
    for field_name, field_spec in field_specs.items():
        value = values[field_name]
        if field_name == "time" and _normalized_output(value.value).upper() in {
            "TBD",
            "TBA",
        }:
            value.source_value = value.value
            value.value = None
            value.missing_reason = "time_pending"
        if _is_empty(value.value):
            issue = _issue_for_missing(field_name, value, field_spec, env)
            if issue:
                issues.append(issue)
        if value.formula_cache_missing:
            formula_policy = runtime.config["workbook"]["formula_value_policy"].get(
                "on_missing_cached_value", "warning"
            )
            if formula_policy == "null":
                continue
            severity = (
                "blocking_error"
                if formula_policy == "blocking_error"
                else "warning"
            )
            normalized_value = _normalized_output(value.value)
            if value.formula_resolution == "resolved_reference":
                message = (
                    "A formula cell had no cached value; its direct cell reference "
                    "was resolved safely."
                )
            elif normalized_value.upper() == "TBD" and (
                value.used_fallback or value.used_default
            ):
                message = (
                    "A formula cell had no cached value; explicit field policy "
                    "used TBD."
                )
            else:
                message = (
                    "A formula cell had no cached value and no safe direct "
                    "reference result was available."
                )
            issues.append(
                ValidationIssue(
                    source_row=anchor.row,
                    source_sheet=runtime.sheet_name(env.source_id),
                    record_set_id=env.record_set_id,
                    severity=severity,
                    code="formula_cached_value_missing",
                    affected_field=field_name,
                    message=message,
                )
            )

    date_text = _normalized_output(values.get("date", _EvalValue()).value)
    time_text = _normalized_output(values.get("time", _EvalValue()).value)
    team_a = _normalized_output(values.get("team_a", _EvalValue()).value)
    team_b = _normalized_output(values.get("team_b", _EvalValue()).value)
    stage = _normalized_output(values.get("stage", _EvalValue()).value)
    bo = _normalized_output(values.get("bo", _EvalValue()).value)
    match_label = _normalized_output(values.get("match_label", _EvalValue()).value)
    bo = _normalize_bo(
        bo,
        bool(runtime.config.get("normalization", {}).get("best_of", {}).get("normalize")),
    )

    if timezone_error:
        issues.append(
            ValidationIssue(
                source_row=anchor.row,
                source_sheet=runtime.sheet_name(env.source_id),
                record_set_id=env.record_set_id,
                severity="blocking_error",
                code=timezone_error,
                affected_field="timezone",
                message=(
                    "ParserKey timezone is required."
                    if timezone_error == "missing_timezone"
                    else f'ParserKey timezone "{runtime.parser_key.base_timezone}" is invalid.'
                ),
            )
        )

    start_time_utc = ""
    if date_text and time_text and zone is not None:
        try:
            start_time_utc = to_utc_string(
                parse_date_value(date_text, epoch=runtime.epoch),
                parse_time_value(time_text, epoch=runtime.epoch),
                zone,
                ambiguous_policy=runtime.config["normalization"]["datetime"].get(
                    "ambiguous_time_policy", "blocking_error"
                ),
            )
        except (ValueError, OverflowError, AmbiguousLocalTimeError, NonexistentLocalTimeError):
            issues.append(
                ValidationIssue(
                    source_row=anchor.row,
                    source_sheet=runtime.sheet_name(env.source_id),
                    record_set_id=env.record_set_id,
                    severity="blocking_error",
                    code="unparseable_datetime",
                    affected_field="start_time_utc",
                    message="The extracted local date and time could not be converted to UTC.",
                )
            )

    provenance = {
        field_name: _provenance(value, env) for field_name, value in values.items()
    }
    match = ParsedMatch(
        source_row=anchor.row,
        source_sheet=runtime.sheet_name(env.source_id),
        record_set_id=env.record_set_id,
        tile_id=anchor.tile_id,
        tile_origin=(
            _source_coordinate(anchor.tile_origin[1], anchor.tile_origin[0])
            if anchor.tile_origin
            else None
        ),
        date_original=date_text,
        time_original=time_text,
        timezone=runtime.parser_key.base_timezone,
        start_time_utc=start_time_utc,
        team_a=team_a,
        team_b=team_b,
        stage=stage,
        bo=bo,
        match_label=match_label,
        field_provenance=provenance,
    )
    return match, issues


def _timezone(parser_key: ParserKey) -> tuple[ZoneInfo | None, str | None]:
    if not parser_key.base_timezone:
        return None, "missing_timezone"
    try:
        return ZoneInfo(parser_key.base_timezone), None
    except (ZoneInfoNotFoundError, ValueError):
        return None, "invalid_timezone"


def _is_placeholder(value: str, team_config: dict[str, Any]) -> bool:
    text = normalize_whitespace(value)
    if not text:
        return False
    placeholder_values = {
        normalize_whitespace(str(item)).casefold()
        for item in team_config.get("placeholder_values", [])
    }
    if text.casefold() in placeholder_values:
        return True
    return any(
        re.fullmatch(pattern, text, re.IGNORECASE) is not None
        for pattern in team_config.get("placeholder_patterns", [])
    )


def _add_duplicates(
    matches: list[ParsedMatch],
    issues: list[ValidationIssue],
    duplicate_config: dict[str, Any],
    team_config: dict[str, Any],
) -> None:
    fields = list(duplicate_config.get("fields", []))
    severity = duplicate_config.get("severity", "warning")
    team_order_sensitive = duplicate_config.get("team_order_sensitive", True)
    groups: dict[tuple[str, ...], list[ParsedMatch]] = defaultdict(list)
    for match in matches:
        if _is_placeholder(match.team_a, team_config) and _is_placeholder(
            match.team_b, team_config
        ):
            continue
        values = {field: str(getattr(match, field, "") or "") for field in fields}
        if not all(values.values()):
            continue
        if not team_order_sensitive and {"team_a", "team_b"}.issubset(values):
            values["team_a"], values["team_b"] = sorted(
                (values["team_a"], values["team_b"]), key=str.casefold
            )
        groups[tuple(values[field] for field in fields)].append(match)
    for group in groups.values():
        if len(group) < 2:
            continue
        coordinates = ", ".join(
            f"{match.source_sheet}!{match.source_row}" for match in group
        )
        for match in group:
            issues.append(
                ValidationIssue(
                    source_row=match.source_row,
                    source_sheet=match.source_sheet,
                    record_set_id=match.record_set_id,
                    severity=severity,
                    code="possible_duplicate",
                    affected_field=",".join(fields),
                    message=f"Possible duplicate across {coordinates}.",
                )
            )


def _assign_statuses(matches: list[ParsedMatch], issues: list[ValidationIssue]) -> None:
    by_record: dict[tuple[str | None, int | None, str | None], list[ValidationIssue]] = defaultdict(list)
    for issue in issues:
        by_record[(issue.source_sheet, issue.source_row, issue.record_set_id)].append(issue)
    for match in matches:
        row_issues = by_record[(match.source_sheet, match.source_row, match.record_set_id)]
        if any(issue.severity == "blocking_error" for issue in row_issues):
            match.row_status = "invalid"
        elif any(issue.severity == "warning" for issue in row_issues):
            match.row_status = "warning"
        else:
            match.row_status = "valid"


def parse_workbook_v2(file_bytes: bytes, parser_key: ParserKey) -> ParseResult:
    notice = (
        "timezone_from_key: match timezone is taken from ParserKey "
        f"({parser_key.base_timezone or '(missing)'})."
    )
    value_workbook = None
    formula_workbook = None
    try:
        value_workbook = load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=False, keep_links=False
        )
        formula_workbook = load_workbook(
            BytesIO(file_bytes), data_only=False, read_only=False, keep_links=False
        )
        runtime = _Runtime(parser_key, value_workbook, formula_workbook)
        zone, timezone_error = _timezone(parser_key)
        matches: list[ParsedMatch] = []
        issues: list[ValidationIssue] = []
        include_hidden_rows = parser_key.raw_data["workbook"][
            "hidden_content_policy"
        ].get("include_hidden_rows", False)

        for record_set in parser_key.raw_data["record_sets"]:
            if not record_set.get("enabled", True):
                continue
            anchors = _anchors(record_set["locator"])
            if not anchors:
                continue
            value_sheet, _ = runtime.sheets(record_set["source_id"])
            if not include_hidden_rows:
                anchors = [
                    anchor
                    for anchor in anchors
                    if not bool(value_sheet.row_dimensions[anchor.row].hidden)
                ]
            if not anchors:
                continue

            context: dict[str, _EvalValue] = {}
            context_env = _Environment(
                runtime=runtime,
                source_id=record_set["source_id"],
                record_set_id=record_set["record_set_id"],
                anchor=anchors[0],
                context=context,
            )
            for key, context_spec in record_set.get("context", {}).items():
                context[key] = _evaluate_pipeline(
                    context_spec["value"], context_env, runtime.epoch
                )

            for anchor in anchors:
                match, record_issues = _parse_record(
                    runtime,
                    record_set,
                    anchor,
                    context,
                    zone,
                    timezone_error,
                )
                if match is not None:
                    matches.append(match)
                    issues.extend(record_issues)

        if not matches:
            return ParseResult.failed(
                "No records were emitted by the ParserKey v2 record sets.", notice=notice
            )

        count_rule = parser_key.raw_data["validation"]["record_count"]
        minimum = count_rule.get("minimum")
        maximum = count_rule.get("maximum")
        below_minimum = minimum is not None and len(matches) < int(minimum)
        above_maximum = maximum is not None and len(matches) > int(maximum)
        if below_minimum or above_maximum:
            expected = (
                f"{minimum}–{maximum}"
                if minimum is not None and maximum is not None
                else f"at least {minimum}"
                if minimum is not None
                else f"at most {maximum}"
            )
            issues.append(
                ValidationIssue(
                    source_row=None,
                    severity=count_rule.get("on_violation", "blocking_error"),
                    code="record_count_mismatch",
                    affected_field=None,
                    message=(
                        f"ParserKey expected {expected} "
                        f"records but emitted {len(matches)}."
                    ),
                )
            )

        duplicate_config = parser_key.raw_data["validation"]["duplicate_check"]
        if duplicate_config.get("enabled", True):
            _add_duplicates(
                matches,
                issues,
                duplicate_config,
                parser_key.raw_data.get("normalization", {}).get("teams", {}),
            )
        _assign_statuses(matches, issues)
        if any(issue.severity == "blocking_error" for issue in issues):
            status = "blocked"
        elif any(issue.severity == "warning" for issue in issues):
            status = "parsed_with_warnings"
        else:
            status = "parsed"
        return ParseResult(
            status=status,
            matches=matches,
            issues=issues,
            notice=notice,
        )
    except KeyError as exc:
        return ParseResult.failed(str(exc).strip("'"), notice=notice)
    except (BadZipFile, InvalidFileException, OSError, EOFError) as exc:
        return ParseResult.failed(
            f"Could not read this XLSX file. Upload a valid .xlsx schedule file. ({exc})",
            notice=notice,
        )
    except Exception as exc:
        return ParseResult.failed(
            f"The ParserKey v2 runtime failed: {type(exc).__name__}: {exc}",
            notice=notice,
        )
    finally:
        if value_workbook is not None:
            value_workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()
