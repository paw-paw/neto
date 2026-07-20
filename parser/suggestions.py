"""Bounded structural ParserKey ranking for XLSX workbooks."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_to_tuple
from openpyxl.utils.exceptions import InvalidFileException

from .models import ParserKey
from .workbook_security import WorkbookSafetyError, validate_xlsx_archive


SAMPLE_ROWS = 32
SAMPLE_COLUMNS = 40
MAX_STRUCTURAL_PROBES = 48
MAX_PROBES_PER_SHEET = 16
MAX_ANCHORS_PER_RECORD_SET = 3
NO_MATCH_SCORE = 45
TOKEN_RE = re.compile(r"[^\W_]+", re.UNICODE)
COORDINATE_RE = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?(\d+)")
EDITION_TOKEN_RE = re.compile(
    r"^(season|series|split|episode|ep|s|sa)(\d+)$", re.IGNORECASE
)
EDITION_MARKERS = {"edition", "ep", "episode", "s", "season", "series", "split"}
STOPWORDS = {
    "a",
    "an",
    "and",
    "best",
    "bracket",
    "complete",
    "doc",
    "event",
    "for",
    "key",
    "main",
    "of",
    "offline",
    "online",
    "players",
    "public",
    "results",
    "schedule",
    "stage",
    "the",
    "tournament",
    "v1",
    "v2",
    "workbook",
    "xlsx",
}
HEADER_HINTS = {
    "date": {"date", "day"},
    "time": {"time", "start", "utc", "local"},
    "team_a": {"team", "team1", "teama", "home", "opponent"},
    "team_b": {"team", "team2", "teamb", "away", "opponent"},
    "stage": {"stage", "phase", "round", "group"},
    "bo": {"bo", "bestof", "format", "maps"},
    "match_label": {"match", "label", "game", "fixture"},
}
CRITICAL_FIELDS = ("date", "time", "team_a", "team_b")
CELL_OPERATORS = {
    "cell.absolute",
    "cell.column",
    "cell.merged_origin",
    "cell.relative",
}


class WorkbookFingerprintError(ValueError):
    """Raised when a workbook cannot be fingerprinted safely."""


@dataclass(frozen=True)
class SheetFingerprint:
    name: str
    max_row: int | None
    max_column: int | None
    sampled_tokens: frozenset[str]
    row_tokens: dict[int, frozenset[str]]
    sampled_cells: dict[tuple[int, int], object]


@dataclass(frozen=True)
class WorkbookFingerprint:
    file_name: str
    sheets: tuple[SheetFingerprint, ...]

    @property
    def sheet_names(self) -> tuple[str, ...]:
        return tuple(sheet.name for sheet in self.sheets)

    @property
    def sampled_tokens(self) -> frozenset[str]:
        return frozenset(
            token for sheet in self.sheets for token in sheet.sampled_tokens
        )


@dataclass(frozen=True)
class ParserKeySuggestion:
    parser_key: ParserKey
    score: int
    confidence: str
    reasons: tuple[str, ...]
    margin: int = 0

    @property
    def recommended(self) -> bool:
        return self.confidence in {"High", "Medium"}


@dataclass(frozen=True)
class _IdentityProfile:
    family: frozenset[str]
    editions: frozenset[str]


@dataclass(frozen=True)
class _SheetExpectation:
    name: str
    required: bool


@dataclass(frozen=True)
class _Anchor:
    sheet_name: str
    record_row: int
    record_column: int
    tile_row: int
    tile_column: int


@dataclass(frozen=True)
class _Probe:
    sheet_name: str
    row: int
    column: int
    field_name: str
    operator: str


@dataclass(frozen=True)
class _Candidate:
    parser_key: ParserKey
    score: int
    reasons: tuple[str, ...]
    all_required_sheets: bool
    sheet_score: int
    probe_score: int
    filename_score: int
    dimension_score: int


def _tokens(value: object) -> set[str]:
    if value is None:
        return set()
    return set(TOKEN_RE.findall(str(value).casefold()))


def _identity_profile(value: object) -> _IdentityProfile:
    family: set[str] = set()
    editions: set[str] = set()
    expect_edition = False
    for token in TOKEN_RE.findall(str(value or "").casefold()):
        combined = EDITION_TOKEN_RE.fullmatch(token)
        if combined:
            marker, number = combined.groups()
            if marker.casefold() == "sa":
                family.add("sa")
            editions.add(number)
            expect_edition = False
            continue
        if token in EDITION_MARKERS:
            expect_edition = True
            continue
        if token.isdigit():
            if not (len(token) == 4 and 1900 <= int(token) <= 2100):
                editions.add(token)
            expect_edition = False
            continue
        if expect_edition:
            expect_edition = False
        if token not in STOPWORDS and len(token) > 1:
            family.add(token)
    return _IdentityProfile(frozenset(family), frozenset(editions))


def _similarity(left: frozenset[str], right: frozenset[str]) -> float:
    if not left or not right:
        return 0.0
    return (2 * len(left & right)) / (len(left) + len(right))


def _reported_dimension(value: object) -> int | None:
    if isinstance(value, int) and not isinstance(value, bool) and value > 0:
        return value
    return None


def fingerprint_workbook(
    file_bytes: bytes, file_name: str = "workbook.xlsx"
) -> WorkbookFingerprint:
    """Read a hard-bounded cell sample plus cheap workbook metadata."""

    try:
        validate_xlsx_archive(file_bytes)
    except WorkbookSafetyError as exc:
        raise WorkbookFingerprintError(
            f"The uploaded XLSX could not be inspected: {exc}"
        ) from exc

    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except (OSError, ValueError, BadZipFile, InvalidFileException) as exc:
        raise WorkbookFingerprintError(
            f"The uploaded XLSX could not be inspected: {exc}"
        ) from exc

    sheets: list[SheetFingerprint] = []
    try:
        for worksheet in workbook.worksheets:
            max_row = _reported_dimension(worksheet.max_row)
            max_column = _reported_dimension(worksheet.max_column)
            row_tokens: dict[int, frozenset[str]] = {}
            sampled_cells: dict[tuple[int, int], object] = {}
            sampled: set[str] = set()
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=SAMPLE_ROWS,
                    min_col=1,
                    max_col=SAMPLE_COLUMNS,
                    values_only=True,
                ),
                start=1,
            ):
                current: set[str] = set()
                for column_number, value in enumerate(row, start=1):
                    current.update(_tokens(value))
                    if value is not None and str(value).strip():
                        sampled_cells[(row_number, column_number)] = value
                row_tokens[row_number] = frozenset(current)
                sampled.update(current)
            sheets.append(
                SheetFingerprint(
                    name=worksheet.title,
                    max_row=max_row,
                    max_column=max_column,
                    sampled_tokens=frozenset(sampled),
                    row_tokens=row_tokens,
                    sampled_cells=sampled_cells,
                )
            )
    except Exception as exc:
        raise WorkbookFingerprintError(
            f"The uploaded XLSX could not be inspected: {exc}"
        ) from exc
    finally:
        workbook.close()
    return WorkbookFingerprint(file_name=file_name, sheets=tuple(sheets))


def _sheet_expectations(parser_key: ParserKey) -> list[_SheetExpectation]:
    if parser_key.schema_version != "neto.parser_key.v2":
        return [_SheetExpectation(parser_key.target_sheet, True)]
    expectations: list[_SheetExpectation] = []
    for source in parser_key.raw_data.get("sources", []):
        if not isinstance(source, dict):
            continue
        locator = source.get("sheet_locator", {})
        args = locator.get("args", {}) if isinstance(locator, dict) else {}
        name = args.get("sheet_name") if isinstance(args, dict) else None
        if isinstance(name, str) and name.strip():
            expectations.append(
                _SheetExpectation(name.strip(), source.get("required", True) is not False)
            )
    deduplicated: dict[str, _SheetExpectation] = {}
    for expectation in expectations:
        key = expectation.name.casefold()
        previous = deduplicated.get(key)
        deduplicated[key] = _SheetExpectation(
            expectation.name,
            expectation.required or bool(previous and previous.required),
        )
    return list(deduplicated.values())


def _source_file_names(parser_key: ParserKey) -> list[str]:
    metadata = parser_key.raw_data.get("metadata", {})
    names: list[str] = []
    if isinstance(metadata, dict):
        legacy = metadata.get("source_schedule_file")
        if isinstance(legacy, str):
            names.append(legacy)
        source_files = metadata.get("source_files", [])
        if isinstance(source_files, list):
            for item in source_files:
                if isinstance(item, dict) and isinstance(item.get("filename"), str):
                    names.append(item["filename"])
    return names


def _sheet_lookup(
    fingerprint: WorkbookFingerprint, expected: str
) -> SheetFingerprint | None:
    exact = next((sheet for sheet in fingerprint.sheets if sheet.name == expected), None)
    if exact:
        return exact
    return next(
        (
            sheet
            for sheet in fingerprint.sheets
            if sheet.name.casefold() == expected.casefold()
        ),
        None,
    )


def _max_coordinate_bounds(value: Any) -> tuple[int, int]:
    max_row = 0
    max_column = 0
    if isinstance(value, dict):
        for name, child in value.items():
            if isinstance(child, int) and not isinstance(child, bool):
                if name in {"row", "start_row", "end_row"}:
                    max_row = max(max_row, child)
            child_row, child_column = _max_coordinate_bounds(child)
            max_row = max(max_row, child_row)
            max_column = max(max_column, child_column)
    elif isinstance(value, list):
        for child in value:
            child_row, child_column = _max_coordinate_bounds(child)
            max_row = max(max_row, child_row)
            max_column = max(max_column, child_column)
    elif isinstance(value, str):
        for coordinate in COORDINATE_RE.findall(value):
            max_column = max(max_column, column_index_from_string(coordinate[0]))
            max_row = max(max_row, int(coordinate[1]))
        if re.fullmatch(r"[A-Z]{1,3}", value):
            max_column = max(max_column, column_index_from_string(value))
    return max_row, max_column


def _expected_dimensions_by_sheet(parser_key: ParserKey) -> dict[str, tuple[int, int]]:
    if parser_key.schema_version != "neto.parser_key.v2":
        columns = [
            column_index_from_string(column)
            for column in parser_key.field_mappings.values()
            if column
        ]
        return {
            parser_key.target_sheet.casefold(): (
                parser_key.data_start_row,
                max(columns, default=1),
            )
        }

    source_sheets: dict[str, str] = {}
    for source in parser_key.raw_data.get("sources", []):
        if not isinstance(source, dict):
            continue
        source_id = source.get("source_id")
        locator = source.get("sheet_locator", {})
        args = locator.get("args", {}) if isinstance(locator, dict) else {}
        sheet_name = args.get("sheet_name") if isinstance(args, dict) else None
        if isinstance(source_id, str) and isinstance(sheet_name, str):
            source_sheets[source_id] = sheet_name

    bounds: dict[str, tuple[int, int]] = {}
    for record_set in parser_key.raw_data.get("record_sets", []):
        if not isinstance(record_set, dict):
            continue
        sheet_name = source_sheets.get(str(record_set.get("source_id") or ""))
        if not sheet_name:
            continue
        row, column = _max_coordinate_bounds(record_set)
        current_row, current_column = bounds.get(sheet_name.casefold(), (0, 0))
        bounds[sheet_name.casefold()] = (
            max(current_row, row),
            max(current_column, column),
        )
    return bounds


def _sheet_score(
    parser_key: ParserKey,
    fingerprint: WorkbookFingerprint,
    frequencies: Counter[str],
    catalog_size: int,
) -> tuple[int, bool, str]:
    expectations = _sheet_expectations(parser_key)
    required = [item for item in expectations if item.required] or expectations
    if not required:
        return 0, False, "No required sheets are declared"

    matched = [item for item in required if _sheet_lookup(fingerprint, item.name)]
    coverage = len(matched) / len(required)
    rarity_values = []
    for item in matched:
        frequency = frequencies[item.name.casefold()]
        rarity_values.append(
            1.0
            if catalog_size <= 1
            else 1.0 - ((frequency - 1) / (catalog_size - 1))
        )
    rarity = sum(rarity_values) / len(rarity_values) if rarity_values else 0.0
    score = round(25 * coverage + 15 * coverage * rarity)
    all_present = len(matched) == len(required)
    if all_present:
        reason = (
            f'Expected sheet "{required[0].name}" matches'
            if len(required) == 1
            else f"All {len(required)} required sheets match"
        )
    elif matched:
        reason = f"{len(matched)}/{len(required)} required sheets match"
    else:
        reason = "Required sheets are missing"
    return score, all_present, reason


def _sample_positions(values: list[int], limit: int) -> list[int]:
    unique = sorted(set(values))
    if len(unique) <= limit:
        return unique
    positions = [unique[0], unique[len(unique) // 2], unique[-1]]
    return list(dict.fromkeys(positions))[:limit]


def _record_anchors(record_set: dict[str, Any], sheet_name: str) -> list[_Anchor]:
    locator = record_set.get("locator", {})
    if not isinstance(locator, dict):
        return []
    operation = locator.get("op")
    args = locator.get("args", {})
    if not isinstance(args, dict):
        return []

    anchors: list[_Anchor] = []
    if operation in {"records.row_scan", "records.row_ranges"}:
        rows: list[int] = []
        if operation == "records.row_scan":
            start = int(args.get("start_row") or 1)
            end = min(int(args.get("end_row") or start), SAMPLE_ROWS)
            step = max(1, int(args.get("step") or 1))
            if start <= end:
                rows.extend(range(start, end + 1, step))
        else:
            for item in args.get("ranges", []):
                if not isinstance(item, dict):
                    continue
                start = int(item.get("start_row") or 1)
                end = min(int(item.get("end_row") or start), SAMPLE_ROWS)
                if start <= end:
                    rows.extend(range(start, end + 1))
        anchor_column = args.get("anchor_column", "A")
        try:
            column = column_index_from_string(str(anchor_column))
        except ValueError:
            return []
        for row in _sample_positions(rows, MAX_ANCHORS_PER_RECORD_SET):
            anchors.append(_Anchor(sheet_name, row, column, row, column))
        return anchors

    if operation == "records.tile_grid":
        record_anchor = args.get("record_anchor", {})
        column_offset = (
            int(record_anchor.get("column_offset") or 0)
            if isinstance(record_anchor, dict)
            else 0
        )
        for tile in args.get("tiles", []):
            if not isinstance(tile, dict) or not isinstance(tile.get("origin"), str):
                continue
            try:
                tile_row, tile_column = coordinate_to_tuple(tile["origin"])
            except ValueError:
                continue
            for row_offset in tile.get("record_row_offsets", [0]):
                if not isinstance(row_offset, int):
                    continue
                record_row = tile_row + row_offset
                record_column = tile_column + column_offset
                if (
                    1 <= record_row <= SAMPLE_ROWS
                    and 1 <= record_column <= SAMPLE_COLUMNS
                ):
                    anchors.append(
                        _Anchor(
                            sheet_name,
                            record_row,
                            record_column,
                            tile_row,
                            tile_column,
                        )
                    )
                if len(anchors) >= MAX_ANCHORS_PER_RECORD_SET:
                    return anchors
    return anchors


def _find_cell_nodes(value: Any) -> list[dict[str, Any]]:
    nodes: list[dict[str, Any]] = []
    if isinstance(value, dict):
        if value.get("op") in CELL_OPERATORS:
            nodes.append(value)
        for child in value.values():
            nodes.extend(_find_cell_nodes(child))
    elif isinstance(value, list):
        for child in value:
            nodes.extend(_find_cell_nodes(child))
    return nodes


def _field_cell_nodes(record_set: dict[str, Any], field_name: str) -> list[dict[str, Any]]:
    fields = record_set.get("fields", {})
    field = fields.get(field_name, {}) if isinstance(fields, dict) else {}
    value = field.get("value", {}) if isinstance(field, dict) else {}
    source = value.get("source", {}) if isinstance(value, dict) else {}
    if not isinstance(source, dict):
        return []
    if source.get("op") in CELL_OPERATORS:
        return [source]
    if source.get("op") == "extractor.output":
        extractor_id = source.get("args", {}).get("extractor_id")
        for extractor in record_set.get("extractors", []):
            if isinstance(extractor, dict) and extractor.get("extractor_id") == extractor_id:
                return _find_cell_nodes(extractor)
    if source.get("op") == "context.get":
        context_key = source.get("args", {}).get("key")
        context = record_set.get("context", {})
        if isinstance(context, dict) and context_key in context:
            return _find_cell_nodes(context[context_key])
    return _find_cell_nodes(source)


def _probe_from_node(
    node: dict[str, Any], anchor: _Anchor, field_name: str
) -> _Probe | None:
    operation = str(node.get("op") or "")
    args = node.get("args", {})
    if not isinstance(args, dict):
        return None
    row = anchor.record_row
    column = anchor.record_column
    if operation in {"cell.column", "cell.merged_origin"}:
        try:
            column = column_index_from_string(str(args.get("column") or ""))
        except ValueError:
            return None
    elif operation == "cell.absolute":
        coordinate = args.get("coordinate") or args.get("cell")
        if not isinstance(coordinate, str):
            return None
        try:
            row, column = coordinate_to_tuple(coordinate)
        except ValueError:
            return None
    elif operation == "cell.relative":
        if args.get("base") == "tile_origin":
            row, column = anchor.tile_row, anchor.tile_column
        row += int(args.get("row_offset") or 0)
        column += int(args.get("column_offset") or 0)
    else:
        return None
    if not (1 <= row <= SAMPLE_ROWS and 1 <= column <= SAMPLE_COLUMNS):
        return None
    return _Probe(anchor.sheet_name, row, column, field_name, operation)


def _v2_probes(parser_key: ParserKey) -> list[_Probe]:
    source_sheets: dict[str, str] = {}
    for source in parser_key.raw_data.get("sources", []):
        if not isinstance(source, dict):
            continue
        locator = source.get("sheet_locator", {})
        args = locator.get("args", {}) if isinstance(locator, dict) else {}
        source_id = source.get("source_id")
        sheet_name = args.get("sheet_name") if isinstance(args, dict) else None
        if isinstance(source_id, str) and isinstance(sheet_name, str):
            source_sheets[source_id] = sheet_name

    probes: list[_Probe] = []
    seen: set[tuple[str, int, int, str]] = set()
    per_sheet: defaultdict[str, int] = defaultdict(int)
    for record_set in parser_key.raw_data.get("record_sets", []):
        if not isinstance(record_set, dict) or record_set.get("enabled", True) is False:
            continue
        sheet_name = source_sheets.get(str(record_set.get("source_id") or ""))
        if not sheet_name:
            continue
        for anchor in _record_anchors(record_set, sheet_name):
            for field_name in CRITICAL_FIELDS:
                for node in _field_cell_nodes(record_set, field_name):
                    probe = _probe_from_node(node, anchor, field_name)
                    if probe is None:
                        continue
                    identity = (
                        probe.sheet_name.casefold(),
                        probe.row,
                        probe.column,
                        probe.field_name,
                    )
                    sheet_key = probe.sheet_name.casefold()
                    if (
                        identity in seen
                        or per_sheet[sheet_key] >= MAX_PROBES_PER_SHEET
                    ):
                        continue
                    seen.add(identity)
                    per_sheet[sheet_key] += 1
                    probes.append(probe)
                    if len(probes) >= MAX_STRUCTURAL_PROBES:
                        return probes
    return probes


def _is_plausible(value: object, field_name: str) -> bool:
    if value is None or not str(value).strip():
        return False
    if field_name in {"team_a", "team_b"}:
        return isinstance(value, str) and bool(_tokens(value))
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip().casefold()
    if field_name == "time":
        return bool(re.search(r"\d{1,2}[:.]\d{2}|\b(?:tbd|noon|midnight)\b", text))
    if field_name == "date":
        return bool(
            re.search(
                r"\d{1,4}[-/.]\d{1,2}|\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)",
                text,
            )
        )
    return True


def _probe_has_value(sheet: SheetFingerprint, probe: _Probe) -> bool:
    value = sheet.sampled_cells.get((probe.row, probe.column))
    if _is_plausible(value, probe.field_name):
        return True
    if probe.operator == "cell.merged_origin":
        for row in range(probe.row - 1, max(0, probe.row - 5), -1):
            value = sheet.sampled_cells.get((row, probe.column))
            if _is_plausible(value, probe.field_name):
                return True
    return False


def _v0_probe_compatibility(
    parser_key: ParserKey, fingerprint: WorkbookFingerprint
) -> tuple[float, int, int]:
    sheet = _sheet_lookup(fingerprint, parser_key.target_sheet)
    if sheet is None:
        return 0.0, 0, 0
    mapped = [field for field, column in parser_key.field_mappings.items() if column]
    header = set(sheet.row_tokens.get(parser_key.header_row, frozenset()))
    header_hits = sum(bool(header & HEADER_HINTS[field]) for field in mapped)
    header_ratio = header_hits / max(1, len(mapped))

    data_hits = 0
    data_total = 0
    for row in range(parser_key.data_start_row, min(SAMPLE_ROWS, parser_key.data_start_row + 4) + 1):
        for field_name in CRITICAL_FIELDS:
            column_name = parser_key.field_mappings.get(field_name)
            if not column_name:
                continue
            column = column_index_from_string(column_name)
            data_total += 1
            if _is_plausible(sheet.sampled_cells.get((row, column)), field_name):
                data_hits += 1
    data_ratio = data_hits / data_total if data_total else 0.0
    ratios = [ratio for ratio, total in ((header_ratio, len(mapped)), (data_ratio, data_total)) if total]
    ratio = sum(ratios) / len(ratios) if ratios else 0.0
    return ratio, header_hits + data_hits, len(mapped) + data_total


def _structural_probe_compatibility(
    parser_key: ParserKey, fingerprint: WorkbookFingerprint
) -> tuple[float, int, int]:
    if parser_key.schema_version != "neto.parser_key.v2":
        return _v0_probe_compatibility(parser_key, fingerprint)
    probes = _v2_probes(parser_key)
    hits = 0
    evaluated = 0
    for probe in probes:
        sheet = _sheet_lookup(fingerprint, probe.sheet_name)
        if sheet is None:
            continue
        evaluated += 1
        if _probe_has_value(sheet, probe):
            hits += 1
    return (hits / evaluated if evaluated else 0.0), hits, evaluated


def _dimension_compatibility(
    parser_key: ParserKey, fingerprint: WorkbookFingerprint
) -> tuple[int, bool]:
    bounds = _expected_dimensions_by_sheet(parser_key)
    known = 0
    covered = 0
    for sheet_name, (required_row, required_column) in bounds.items():
        sheet = _sheet_lookup(fingerprint, sheet_name)
        if sheet is None or sheet.max_row is None or sheet.max_column is None:
            continue
        known += 1
        if sheet.max_row >= required_row and sheet.max_column >= required_column:
            covered += 1
    if not known:
        return 0, False
    return round(5 * covered / known), covered == known


def _filename_compatibility(
    parser_key: ParserKey, upload_profile: _IdentityProfile
) -> tuple[int, str | None]:
    values = [
        *_source_file_names(parser_key),
        parser_key.key_name,
        parser_key.tournament_name,
    ]
    best_similarity = 0.0
    best_profile = _IdentityProfile(frozenset(), frozenset())
    for value in values:
        profile = _identity_profile(value)
        similarity = _similarity(upload_profile.family, profile.family)
        if similarity > best_similarity:
            best_similarity = similarity
            best_profile = profile
    score = round(20 * best_similarity)
    if best_similarity < 0.25:
        return score, None
    editions_differ = bool(
        best_similarity >= 0.65
        and upload_profile.editions
        and best_profile.editions
        and upload_profile.editions.isdisjoint(best_profile.editions)
    )
    if editions_differ:
        return score, "Same tournament family; ParserKey is for a different edition"
    if best_similarity >= 0.75:
        return score, "Workbook name matches the tournament family"
    return score, "Workbook name partially matches the tournament family"


def _confidence(candidate: _Candidate, margin: int) -> str:
    evidence_groups = sum(
        (
            candidate.sheet_score >= 20,
            candidate.probe_score >= 10,
            candidate.filename_score >= 8,
            candidate.dimension_score >= 3,
        )
    )
    has_structural_evidence = (
        candidate.probe_score >= 8 or candidate.sheet_score >= 38
    )
    if (
        candidate.score >= 80
        and candidate.all_required_sheets
        and evidence_groups >= 2
        and has_structural_evidence
        and margin >= 12
    ):
        return "High"
    if (
        candidate.score >= 60
        and candidate.all_required_sheets
        and evidence_groups >= 2
        and has_structural_evidence
        and margin >= 8
    ):
        return "Medium"
    return "Low"


def rank_parser_keys(
    fingerprint: WorkbookFingerprint,
    parser_keys: Iterable[ParserKey],
    *,
    limit: int = 3,
) -> list[ParserKeySuggestion]:
    """Rank keys with bounded structural probes and no complete parser execution."""

    keys = list(parser_keys)
    if not keys or limit <= 0:
        return []
    frequencies: Counter[str] = Counter(
        expectation.name.casefold()
        for parser_key in keys
        for expectation in {
            item.name.casefold(): item for item in _sheet_expectations(parser_key)
        }.values()
        if expectation.required
    )
    upload_profile = _identity_profile(Path(fingerprint.file_name).stem)
    candidates: list[_Candidate] = []

    for parser_key in keys:
        sheet_score, all_present, sheet_reason = _sheet_score(
            parser_key, fingerprint, frequencies, len(keys)
        )
        probe_ratio, probe_hits, probe_total = _structural_probe_compatibility(
            parser_key, fingerprint
        )
        probe_score = round(35 * probe_ratio)
        filename_score, filename_reason = _filename_compatibility(
            parser_key, upload_profile
        )
        dimension_score, dimensions_cover = _dimension_compatibility(
            parser_key, fingerprint
        )
        score = min(100, sheet_score + probe_score + filename_score + dimension_score)
        if score <= 0:
            continue

        reasons = [sheet_reason]
        if probe_total and probe_hits:
            reasons.append(f"{probe_hits}/{probe_total} structural probes match")
        if filename_reason:
            reasons.append(filename_reason)
        if dimensions_cover:
            reasons.append("Known sheet bounds cover the ParserKey structure")
        candidates.append(
            _Candidate(
                parser_key=parser_key,
                score=score,
                reasons=tuple(reasons[:3]),
                all_required_sheets=all_present,
                sheet_score=sheet_score,
                probe_score=probe_score,
                filename_score=filename_score,
                dimension_score=dimension_score,
            )
        )

    candidates.sort(
        key=lambda candidate: (
            -candidate.score,
            candidate.parser_key.key_name.casefold(),
            candidate.parser_key.parser_key_id.casefold(),
        )
    )
    if not candidates or candidates[0].score < NO_MATCH_SCORE:
        return []

    suggestions: list[ParserKeySuggestion] = []
    selected = candidates[:limit]
    for index, candidate in enumerate(selected):
        runner_up_score = candidates[index + 1].score if index + 1 < len(candidates) else 0
        margin = max(0, candidate.score - runner_up_score)
        suggestions.append(
            ParserKeySuggestion(
                parser_key=candidate.parser_key,
                score=candidate.score,
                confidence=_confidence(candidate, margin),
                reasons=candidate.reasons,
                margin=margin,
            )
        )
    return suggestions
